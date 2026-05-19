"""Parser scaffolding tests.

Functional per-sheet extraction tests come in PR #63 onward; this
file covers the scaffolding: public API surface, model construction,
validator integration seam, exception hierarchy, CLI smoke.
"""

from datetime import datetime, timezone
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from mfip.ingestion.bloomberg.exceptions import (
    BloombergParserError,
    WorkbookExtractionError,
    WorkbookValidationError,
)
from mfip.ingestion.bloomberg.models import (
    AnalystRecommendationsSheet,
    BetaSheet,
    DividendSheet,
    EarningsEstimatesSheet,
    ParsedCompanyData,
    PriceHistorySheet,
    RVCompsSheet,
)
from mfip.ingestion.bloomberg.parser import (
    _NA_VALUES,
    _extract_beta,
    _extract_ee,
    _extract_hp_daily,
    _extract_hp_monthly,
    _read_price_series,
    _read_scalar,
    ValidationPolicy,
    parse_fx_workbook,
    parse_indices_workbook,
    parse_workbook,
)
from mfip.ingestion.bloomberg.validator import ValidationReport

REPO_ROOT = Path(__file__).resolve().parents[3]
TEMPLATES = REPO_ROOT / "templates" / "bloomberg"
MASTER_TEMPLATE = TEMPLATES / "Template_Ticker" / "MFIP_Bloomberg_Export_Template_Master.xlsx"
INDICES_TEMPLATE = TEMPLATES / "Template_Index" / "MFIP_Bloomberg_Export_Template_Indices.xlsx"
FX_TEMPLATE = TEMPLATES / "Template_FX" / "MFIP_Bloomberg_Export_Template_FX.xlsx"

FIXTURES = Path(__file__).resolve().parent.parent.parent / "fixtures" / "bloomberg"
EQNR_FIXTURE = FIXTURES / "EQNR_NO_2026-05-08.xlsx"


# ---------------------------------------------------------------------------
# Public surface importability
# ---------------------------------------------------------------------------

class TestPublicSurface:
    def test_parser_module_exports(self):
        """The public names are importable and have expected shapes."""
        assert callable(parse_workbook)
        assert callable(parse_indices_workbook)
        assert callable(parse_fx_workbook)
        assert ValidationPolicy.STRICT.value == "strict"
        assert ValidationPolicy.PERMISSIVE.value == "permissive"
        # REPORT_ONLY intentionally dropped — verify it's not there
        assert not hasattr(ValidationPolicy, "REPORT_ONLY")

    def test_exception_hierarchy(self):
        assert issubclass(WorkbookValidationError, BloombergParserError)
        assert issubclass(WorkbookExtractionError, BloombergParserError)


# ---------------------------------------------------------------------------
# Model construction
# ---------------------------------------------------------------------------

class TestModels:
    def test_parsed_company_data_default_construction(self):
        """ParsedCompanyData with no extraction populated still validates."""
        # ValidationReport's dataclass takes `file=Path` (NOT path=).
        report = ValidationReport(file=Path("dummy.xlsx"))

        data = ParsedCompanyData(
            source_path=Path("dummy.xlsx"),
            parsed_at=datetime.now(timezone.utc),
            validation_report=report,
        )
        assert data.ticker == ""
        assert data.config_currency == ""
        assert data.rv_comps is None
        assert isinstance(data.beta, BetaSheet)
        assert isinstance(data.dvd, DividendSheet)
        assert data.is_gbp_pence is False

    def test_sub_models_default_empty(self):
        """All sheet sub-models default to empty without errors."""
        assert BetaSheet().raw_beta is None
        assert DividendSheet().records == []
        assert EarningsEstimatesSheet().best_eps is None
        assert AnalystRecommendationsSheet().broker_recommendations == []
        assert PriceHistorySheet().series is None
        assert RVCompsSheet().peers == []

    def test_config_currency_accepts_any_string(self):
        """config_currency is plain str — non-USD values are valid at parse
        time (validator's job to ADVISE, not parser's job to refuse)."""
        report = ValidationReport(file=Path("dummy.xlsx"))
        data = ParsedCompanyData(
            source_path=Path("dummy.xlsx"),
            parsed_at=datetime.now(timezone.utc),
            validation_report=report,
            config_currency="NOK",
        )
        assert data.config_currency == "NOK"


# ---------------------------------------------------------------------------
# Validator-integration seam
# ---------------------------------------------------------------------------

class TestValidatorSeam:
    def test_parse_master_template_permissive_raises_on_config(self):
        """Master template has B3=None. PERMISSIVE proceeds past
        validator FAIL but CONFIG extraction raises."""
        with pytest.raises(WorkbookExtractionError, match="CONFIG"):
            parse_workbook(
                MASTER_TEMPLATE,
                validation_policy=ValidationPolicy.PERMISSIVE,
            )

    def test_parse_nonexistent_strict_raises_validation_error(self):
        """The validator catches missing files (FILE_NOT_FOUND code)
        and returns FAIL — STRICT policy promotes to
        WorkbookValidationError before extraction runs."""
        with pytest.raises(WorkbookValidationError):
            parse_workbook(Path("nonexistent.xlsx"))


# ---------------------------------------------------------------------------
# Sibling parser skeletons
# ---------------------------------------------------------------------------

class TestIndicesAndFX:
    def test_parse_indices_skeleton(self):
        """parse_indices_workbook returns a populated-but-empty model."""
        result = parse_indices_workbook(INDICES_TEMPLATE)
        assert result.source_path == INDICES_TEMPLATE
        assert result.obx is None
        assert result.spx is None

    def test_parse_fx_skeleton(self):
        result = parse_fx_workbook(FX_TEMPLATE)
        assert result.source_path == FX_TEMPLATE
        assert result.nokusd is None


# ---------------------------------------------------------------------------
# _read_scalar helper — unit tests against in-memory workbooks
# ---------------------------------------------------------------------------

def _make_ws(title: str, values: dict[str, object]) -> openpyxl.worksheet.worksheet.Worksheet:
    """Build a worksheet with `title` and the given coord -> value cells."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    for coord, value in values.items():
        ws[coord] = value
    return ws


class TestReadScalar:
    @pytest.mark.parametrize("na_value", sorted(_NA_VALUES))
    def test_handles_each_known_na_variant(self, na_value: str):
        """Every enumerated _NA_VALUES entry returns None and is recorded."""
        ws = _make_ws("BETA", {"C6": na_value})
        na_cells: list[str] = []
        assert _read_scalar(ws, "C6", na_cells=na_cells) is None
        assert na_cells == ["BETA!C6"]

    def test_handles_unknown_hash_prefix(self):
        """Future Excel error codes are caught by the '#' prefix guard."""
        ws = _make_ws("EE", {"C8": "#FUTURE_ERROR_CODE!"})
        na_cells: list[str] = []
        assert _read_scalar(ws, "C8", na_cells=na_cells) is None
        assert na_cells == ["EE!C8"]

    def test_handles_none_cell(self):
        ws = _make_ws("BETA", {})  # C6 unset → None
        na_cells: list[str] = []
        assert _read_scalar(ws, "C6", na_cells=na_cells) is None
        assert na_cells == ["BETA!C6"]

    def test_handles_empty_string(self):
        ws = _make_ws("EE", {"C7": ""})
        na_cells: list[str] = []
        assert _read_scalar(ws, "C7", na_cells=na_cells) is None
        assert na_cells == ["EE!C7"]

    def test_handles_numeric_string_with_comma(self):
        """Bloomberg occasionally emits comma-grouped numbers; parse them."""
        ws = _make_ws("EE", {"C9": "1,234.5"})
        na_cells: list[str] = []
        assert _read_scalar(ws, "C9", na_cells=na_cells) == 1234.5
        assert na_cells == []

    def test_raises_on_unexpected_string(self):
        ws = _make_ws("EE", {"C8": "approx 1.2"})
        with pytest.raises(WorkbookExtractionError) as excinfo:
            _read_scalar(ws, "C8", na_cells=[])
        assert "EE!C8" in str(excinfo.value)
        assert "approx 1.2" in str(excinfo.value)

    def test_handles_int_value(self):
        ws = _make_ws("BETA", {"C6": 2})
        na_cells: list[str] = []
        result = _read_scalar(ws, "C6", na_cells=na_cells)
        assert result == 2.0
        assert isinstance(result, float)
        assert na_cells == []

    def test_handles_float_value(self):
        ws = _make_ws("BETA", {"C6": 1.524651})
        na_cells: list[str] = []
        assert _read_scalar(ws, "C6", na_cells=na_cells) == 1.524651
        assert na_cells == []


# ---------------------------------------------------------------------------
# Tier-1: BETA/EE extraction against the master template
#
# The template was saved without the Bloomberg add-in active, so cached
# BDP formula values are Excel "#NAME?" errors. data_only=True returns
# those as strings, which _read_scalar treats as NA-likes (via the '#'
# prefix guard). Result: all-None models with full na_cells coverage.
# ---------------------------------------------------------------------------

class TestExtractTemplate:
    def test_extract_beta_template_all_na(self):
        wb = openpyxl.load_workbook(MASTER_TEMPLATE, data_only=True, read_only=True)
        try:
            beta = _extract_beta(wb)
        finally:
            wb.close()
        assert beta.raw_beta is None
        assert beta.adjusted_beta is None
        assert beta.r_squared is None
        assert beta.na_cells == ["BETA!C6", "BETA!C7", "BETA!C8"]

    def test_extract_ee_template_all_na(self):
        wb = openpyxl.load_workbook(MASTER_TEMPLATE, data_only=True, read_only=True)
        try:
            ee = _extract_ee(wb)
        finally:
            wb.close()
        assert ee.best_eps is None
        assert ee.best_ltg_eeps is None
        assert ee.best_sales is None
        assert ee.best_ebitda is None
        assert ee.best_analyst_rating is None
        assert ee.na_cells == [
            "EE!C10", "EE!C11", "EE!C7", "EE!C8", "EE!C9",
        ]  # sorted lexicographically — matches the parser's sorted() at assembly


# ---------------------------------------------------------------------------
# Tier-2: BETA/EE extraction against the committed real-export fixture.
# EQNR's 2026-05-08 export has values for every scalar in BETA and EE
# (no #N/A). Tests assert types and plausible ranges, not exact numbers —
# Bloomberg cached values drift on re-save.
# ---------------------------------------------------------------------------

class TestExtractEqnrFixture:
    def test_extract_beta_eqnr_real_values(self):
        wb = openpyxl.load_workbook(EQNR_FIXTURE, data_only=True, read_only=True)
        try:
            beta = _extract_beta(wb)
        finally:
            wb.close()
        assert isinstance(beta.raw_beta, float)
        assert isinstance(beta.adjusted_beta, float)
        assert isinstance(beta.r_squared, float)
        assert 0 < beta.raw_beta < 3
        assert 0 < beta.adjusted_beta < 3
        assert 0 <= beta.r_squared <= 1
        assert beta.na_cells == []

    def test_extract_ee_eqnr_real_values(self):
        wb = openpyxl.load_workbook(EQNR_FIXTURE, data_only=True, read_only=True)
        try:
            ee = _extract_ee(wb)
        finally:
            wb.close()
        for field in (
            ee.best_eps,
            ee.best_ltg_eeps,
            ee.best_sales,
            ee.best_ebitda,
            ee.best_analyst_rating,
        ):
            assert isinstance(field, float)
        assert ee.best_eps > 0
        # best_ltg_eeps is a decimal (0.05 = 5%) due to the template's
        # `%` formula suffix; sanity-check it's not accidentally a percent.
        assert 0 < ee.best_ltg_eeps < 0.5
        assert ee.best_sales > 0
        assert ee.best_ebitda > 0
        assert 1.0 <= ee.best_analyst_rating <= 5.0
        assert ee.na_cells == []

    def test_parse_eqnr_full_smoke(self):
        """End-to-end: STRICT parse of the fixture populates BETA + EE,
        validator reports PASS or ADVISORY (not FAIL), and parser does
        not raise."""
        result = parse_workbook(EQNR_FIXTURE)
        assert result.ticker_short == "EQNR_NO"
        assert result.beta.raw_beta is not None
        assert result.ee.best_eps is not None
        assert result.beta.na_cells == []
        assert result.ee.na_cells == []


# ---------------------------------------------------------------------------
# _read_price_series helper — synthetic in-memory workbook unit tests
# ---------------------------------------------------------------------------

def _make_hp_ws(
    title: str,
    rows: list[tuple[object, object]],
    *,
    headers: tuple[str, str] = ("Date", "PX_LAST"),
) -> openpyxl.worksheet.worksheet.Worksheet:
    """Build a price-history-shaped worksheet.

    Rows go into B6/C6 onward. None values are written as None so the
    cell stays empty (mirrors the spilled-but-empty case in templates).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    ws["B5"] = headers[0]
    ws["C5"] = headers[1]
    for i, (b, c) in enumerate(rows):
        ws.cell(row=6 + i, column=2, value=b)
        ws.cell(row=6 + i, column=3, value=c)
    return ws


class TestReadPriceSeries:
    def test_terminates_at_first_empty_row(self):
        rows = [
            (datetime(2025, 1, 31), 100.0),
            (datetime(2025, 2, 28), 101.0),
            (datetime(2025, 3, 31), 102.0),
            (None, None),  # End of spill at row 9.
            (datetime(2025, 5, 31), 999.0),  # Must NOT be picked up.
        ]
        ws = _make_hp_ws("HP_Monthly", rows)
        na_cells: list[str] = []
        series = _read_price_series(ws, na_cells=na_cells)
        assert len(series) == 3
        assert list(series.values) == [100.0, 101.0, 102.0]
        assert na_cells == []

    def test_terminates_on_hash_prefix_in_date_col(self):
        """Template-state case: B6 holds '#NAME?' as a string."""
        rows = [("#NAME?", None)]
        ws = _make_hp_ws("HP_Monthly", rows)
        na_cells: list[str] = []
        series = _read_price_series(ws, na_cells=na_cells)
        assert len(series) == 0
        assert na_cells == ["HP_Monthly!B6"]

    def test_handles_hash_prefix_in_value_col(self):
        """Value sentinel becomes NaN; coordinate recorded."""
        rows = [
            (datetime(2025, 1, 31), 100.0),
            (datetime(2025, 2, 28), "#N/A"),
            (datetime(2025, 3, 31), 102.0),
        ]
        ws = _make_hp_ws("HP_Monthly", rows)
        na_cells: list[str] = []
        series = _read_price_series(ws, na_cells=na_cells)
        assert len(series) == 3
        assert series.iloc[0] == 100.0
        assert pd.isna(series.iloc[1])
        assert series.iloc[2] == 102.0
        assert na_cells == ["HP_Monthly!C7"]

    def test_handles_none_value_with_date_present(self):
        """A date with a None price is kept with NaN, not skipped."""
        rows = [
            (datetime(2025, 1, 31), 100.0),
            (datetime(2025, 2, 28), None),
            # No trailing empty row needed — termination requires both
            # B and C None on the SAME row. Use an explicit end marker.
            (datetime(2025, 3, 31), 102.0),
            (None, None),
        ]
        ws = _make_hp_ws("HP_Monthly", rows)
        na_cells: list[str] = []
        series = _read_price_series(ws, na_cells=na_cells)
        assert len(series) == 3
        assert pd.isna(series.iloc[1])
        assert na_cells == ["HP_Monthly!C7"]

    def test_skips_row_with_none_date(self):
        """Price without a date is unusable; row is skipped entirely."""
        rows = [
            (datetime(2025, 1, 31), 100.0),
            (None, 999.0),  # Date-less price — skip.
            (datetime(2025, 3, 31), 102.0),
        ]
        ws = _make_hp_ws("HP_Monthly", rows)
        na_cells: list[str] = []
        series = _read_price_series(ws, na_cells=na_cells)
        assert len(series) == 2
        assert list(series.values) == [100.0, 102.0]
        assert na_cells == ["HP_Monthly!B7"]

    def test_raises_on_non_date_in_date_col(self):
        """Non-`#` strings in the date column are workbook corruption."""
        rows = [("not a date", 100.0)]
        ws = _make_hp_ws("HP_Monthly", rows)
        with pytest.raises(WorkbookExtractionError) as excinfo:
            _read_price_series(ws, na_cells=[])
        assert "HP_Monthly!B6" in str(excinfo.value)
        assert "not a date" in str(excinfo.value)

    def test_raises_on_unexpected_string_in_value_col(self):
        rows = [(datetime(2025, 1, 31), "approx 1.2")]
        ws = _make_hp_ws("HP_Monthly", rows)
        with pytest.raises(WorkbookExtractionError) as excinfo:
            _read_price_series(ws, na_cells=[])
        assert "HP_Monthly!C6" in str(excinfo.value)
        assert "approx 1.2" in str(excinfo.value)

    def test_handles_numeric_string_with_comma(self):
        rows = [(datetime(2025, 1, 31), "1,234.5")]
        ws = _make_hp_ws("HP_Monthly", rows)
        na_cells: list[str] = []
        series = _read_price_series(ws, na_cells=na_cells)
        assert len(series) == 1
        assert series.iloc[0] == 1234.5
        assert na_cells == []

    def test_returned_series_has_px_last_name_and_datetime_index(self):
        rows = [(datetime(2025, 1, 31), 100.0), (datetime(2025, 2, 28), 101.0)]
        ws = _make_hp_ws("HP_Monthly", rows)
        series = _read_price_series(ws, na_cells=[])
        assert series.name == "PX_LAST"
        assert isinstance(series.index, pd.DatetimeIndex)
        assert series.index.name == "Date"
        assert series.dtype == "float64"


# ---------------------------------------------------------------------------
# Tier 1 — HP extraction against the master template
#
# Template has B6 = '#NAME?' in the date column for both HP sheets.
# Extractor returns an empty series with one na_cells entry pointing
# to B6 (the symmetric '#'-prefix guard from PR #65).
# ---------------------------------------------------------------------------

class TestExtractHpTemplate:
    def test_extract_hp_monthly_template_empty(self):
        wb = openpyxl.load_workbook(MASTER_TEMPLATE, data_only=True, read_only=True)
        try:
            hp = _extract_hp_monthly(wb)
        finally:
            wb.close()
        assert len(hp.series) == 0
        assert hp.series.name == "PX_LAST"
        assert isinstance(hp.series.index, pd.DatetimeIndex)
        assert hp.na_cells == ["HP_Monthly!B6"]

    def test_extract_hp_daily_template_empty(self):
        wb = openpyxl.load_workbook(MASTER_TEMPLATE, data_only=True, read_only=True)
        try:
            hp = _extract_hp_daily(wb)
        finally:
            wb.close()
        assert len(hp.series) == 0
        assert hp.na_cells == ["HP_Daily!B6"]

    def test_extract_hp_monthly_raises_on_header_mismatch(self):
        """Cheap safety net for accidental row-shift before save."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "HP_Monthly"
        ws["B5"] = "Wrong"
        ws["C5"] = "PX_LAST"
        # Wrap in a fake-workbook-like dict: parser uses wb["HP_Monthly"].
        class FakeWb:
            def __getitem__(self, key):
                assert key == "HP_Monthly"
                return ws
        with pytest.raises(WorkbookExtractionError, match="header mismatch"):
            _extract_hp_monthly(FakeWb())


# ---------------------------------------------------------------------------
# Tier 2 — HP extraction against the committed EQNR fixture (PR #64)
# ---------------------------------------------------------------------------

class TestExtractHpEqnrFixture:
    def test_hp_monthly_eqnr_shape_and_range(self):
        wb = openpyxl.load_workbook(EQNR_FIXTURE, data_only=True, read_only=True)
        try:
            hp = _extract_hp_monthly(wb)
        finally:
            wb.close()
        assert 100 <= len(hp.series) <= 130   # 10y * 12m + tolerance
        assert hp.series.name == "PX_LAST"
        assert isinstance(hp.series.index, pd.DatetimeIndex)
        assert hp.series.dtype == "float64"
        assert (hp.series.dropna() > 0).all()
        # Sane window around 2026-05-08 export date.
        assert hp.series.index.min() >= pd.Timestamp("2015-01-01")
        assert hp.series.index.max() <= pd.Timestamp("2026-06-01")
        assert hp.na_cells == []

    def test_hp_daily_eqnr_shape_and_range(self):
        wb = openpyxl.load_workbook(EQNR_FIXTURE, data_only=True, read_only=True)
        try:
            hp = _extract_hp_daily(wb)
        finally:
            wb.close()
        assert 1100 <= len(hp.series) <= 1400  # ~5y * 252 trading days
        assert hp.series.name == "PX_LAST"
        assert isinstance(hp.series.index, pd.DatetimeIndex)
        assert hp.series.dtype == "float64"
        assert (hp.series.dropna() > 0).all()
        assert hp.series.index.min() >= pd.Timestamp("2020-01-01")
        assert hp.series.index.max() <= pd.Timestamp("2026-06-01")
        assert hp.na_cells == []

    def test_parse_eqnr_full_smoke_includes_hp(self):
        result = parse_workbook(EQNR_FIXTURE)
        assert result.hp_monthly.series is not None
        assert result.hp_daily.series is not None
        assert len(result.hp_monthly.series) > 0
        assert len(result.hp_daily.series) > 0
        assert result.hp_monthly.na_cells == []
        assert result.hp_daily.na_cells == []
