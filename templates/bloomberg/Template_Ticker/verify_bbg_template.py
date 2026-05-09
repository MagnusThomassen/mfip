from openpyxl import load_workbook

wb = load_workbook(r"C:\Users\MagnusThomassen\MFIP_Bloomberg_Export_Template_Master.xlsx")
print("Sheets:", wb.sheetnames)
print()

cfg = wb["CONFIG"]
print(f"CONFIG B3 (TICKER input):   {cfg['B3'].value!r}   fill={cfg['B3'].fill.start_color.rgb if cfg['B3'].fill else None}")
print(f"CONFIG B4 (CURRENCY input): {cfg['B4'].value!r}   fill={cfg['B4'].fill.start_color.rgb if cfg['B4'].fill else None}")
print()

for name in ["BETA", "DVD", "EE", "RV_Comps", "ANR", "HP_Monthly", "HP_Daily"]:
    ws = wb[name]
    print(f"=== {name} ===")
    formulas_found = 0
    refs_config = 0
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.startswith("="):
                formulas_found += 1
                if "CONFIG!$B$3" in v or "CONFIG!$B$4" in v:
                    refs_config += 1
                if formulas_found <= 6:
                    print(f"  {cell.coordinate}: {v}")
    print(f"  -> {formulas_found} formula(s), {refs_config} reference CONFIG!$B$3 or $B$4")
    print()
