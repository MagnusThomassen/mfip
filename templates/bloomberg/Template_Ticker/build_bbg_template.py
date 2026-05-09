from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ---------------- Style helpers ----------------
ARIAL = "Arial"
font_title = Font(name=ARIAL, size=14, bold=True, color="FFFFFF")
font_section = Font(name=ARIAL, size=11, bold=True, color="FFFFFF")
font_label = Font(name=ARIAL, size=10, bold=True)
font_input = Font(name=ARIAL, size=11, bold=True, color="000000")
font_body = Font(name=ARIAL, size=10)
font_formula = Font(name=ARIAL, size=10, color="000000")  # black = formula
font_note = Font(name=ARIAL, size=9, italic=True, color="595959")

fill_title = PatternFill("solid", start_color="002060")     # navy
fill_section = PatternFill("solid", start_color="305496")   # blue
fill_input = PatternFill("solid", start_color="F8CBAD")     # light red / salmon
fill_header = PatternFill("solid", start_color="D9E1F2")    # light blue
fill_note = PatternFill("solid", start_color="FFF2CC")      # light yellow

thin = Side(border_style="thin", color="BFBFBF")
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_header_row(ws, row, cols, fill=fill_header, font=font_label):
    for c in cols:
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = center
        cell.border = border_all


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------------- CONFIG sheet ----------------
ws = wb.active
ws.title = "CONFIG"
set_widths(ws, [4, 28, 38, 22, 22, 18])

ws.merge_cells("B1:F1")
c = ws["B1"]
c.value = "MFIP Bloomberg Export Template — Master"
c.font = font_title
c.fill = fill_title
c.alignment = center
ws.row_dimensions[1].height = 26

# Inputs (highlighted light red)
ws["B3"] = "TICKER"
ws["B3"].font = font_label
ws["B3"].alignment = left
ws["C3"] = "EQNR NO Equity"
ws["C3"].font = font_input
ws["C3"].fill = fill_input
ws["C3"].alignment = center
ws["C3"].border = border_all

ws["B4"] = "CURRENCY"
ws["B4"].font = font_label
ws["B4"].alignment = left
ws["C4"] = "USD"
ws["C4"].font = font_input
ws["C4"].fill = fill_input
ws["C4"].alignment = center
ws["C4"].border = border_all

# NOTE: spec says inputs in B3/B4. We keep the label in B and the input value in C
# but ALSO mirror the value into B3/B4 per spec wording. Spec literally says
# "B3: TICKER (default: 'EQNR NO Equity')" — interpreting this as: cell B3 holds
# the ticker value. We use B3/B4 as the live input cells and reference them in formulas.
# Re-do: place inputs DIRECTLY in B3 and B4 to match spec exactly.

# Reset and redo so B3 = ticker value, B4 = currency value (per spec)
ws["B3"] = "EQNR NO Equity"
ws["B4"] = "USD"
for cell_ref in ("B3", "B4"):
    cc = ws[cell_ref]
    cc.font = font_input
    cc.fill = fill_input
    cc.alignment = center
    cc.border = border_all

# Labels to the left (column A is narrow, so put labels in C)
ws["C3"] = "  <-- TICKER (change me)"
ws["C3"].font = Font(name=ARIAL, size=10, italic=True, color="C00000")
ws["C3"].alignment = left
ws["C3"].fill = PatternFill(fill_type=None)
ws["C3"].border = Border()

ws["C4"] = "  <-- CURRENCY"
ws["C4"].font = Font(name=ARIAL, size=10, italic=True, color="C00000")
ws["C4"].alignment = left
ws["C4"].fill = PatternFill(fill_type=None)
ws["C4"].border = Border()

# Instructions block
ws.merge_cells("B6:F6")
c = ws["B6"]
c.value = "Lab Workflow — Instructions"
c.font = font_section
c.fill = fill_section
c.alignment = center

instructions = [
    "1. Open this template on a Bloomberg Terminal lab PC (Bloomberg Excel Add-in must be loaded).",
    "2. Change the ticker in cell B3 to your desired security (e.g., \"DNB NO Equity\", \"AAPL US Equity\").",
    "3. Optionally change the reporting currency in B4 (USD, NOK, EUR, ...).",
    "4. Wait ~30 seconds for all BDP() and BDH() calls across the seven data sheets to refresh.",
    "5. Verify no #N/A Requesting Data... cells remain (press F9 if needed).",
    "6. Select all (Ctrl+A) on each data sheet, Copy, then Paste Special > Values to freeze the data.",
    "7. Save As: <TICKER>_<YYYY-MM-DD>.xlsx (e.g., DNB_2026-05-08.xlsx).",
    "8. Done — the static export is ready for the MFIP report.",
]
for i, txt in enumerate(instructions, start=7):
    ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=6)
    cell = ws.cell(row=i, column=2, value=txt)
    cell.font = font_body
    cell.alignment = left
    ws.row_dimensions[i].height = 20

# Quick reference table
qref_start = 7 + len(instructions) + 1  # blank row then header
ws.merge_cells(start_row=qref_start, start_column=2, end_row=qref_start, end_column=6)
c = ws.cell(row=qref_start, column=2, value="Quick Reference — Key Metrics Extracted")
c.font = font_section
c.fill = fill_section
c.alignment = center

headers = ["Sheet", "Data Type", "Bloomberg Function", "Key Fields", "Frequency"]
hr = qref_start + 1
for j, h in enumerate(headers, start=2):
    cell = ws.cell(row=hr, column=j, value=h)
    cell.font = font_label
    cell.fill = fill_header
    cell.alignment = center
    cell.border = border_all

qref_rows = [
    ("BETA",       "Beta & R-squared",      "BDP",        "BETA_RAW_OVERRIDABLE, BETA_ADJ_OVERRIDABLE, R_SQR",                "Point-in-time"),
    ("DVD",        "Dividend history",      "BDS",        "DVD_HIST_ALL (Decl, Ex, Record, Pay, Amt, Type, Ccy)",             "All history"),
    ("EE",         "Earnings estimates",    "BDP",        "BEST_EPS, BEST_LTG_EPS, BEST_SALES, BEST_EBITDA, BEST_ANALYST_RATING", "Forward consensus"),
    ("RV_Comps",   "Relative valuation",    "BDP",        "CURR_ENTP_VAL/BEST_EBITDA, BEST_PE_RATIO, EV/Sales, PX_TO_BOOK_RATIO", "Point-in-time"),
    ("ANR",        "Analyst recommendations","BDP",       "TOT_BUY_REC, TOT_HOLD_REC, TOT_SELL_REC, BEST_TARGET_PRICE, TOT_ANALYST_REC", "Point-in-time"),
    ("HP_Monthly", "Price history",         "BDH",        "PX_LAST (Per=cm)",                                                  "10Y monthly"),
    ("HP_Daily",   "Price history",         "BDH",        "PX_LAST (Per=cd)",                                                  "5Y daily"),
]
for i, row in enumerate(qref_rows, start=hr + 1):
    for j, val in enumerate(row, start=2):
        cell = ws.cell(row=i, column=j, value=val)
        cell.font = font_body
        cell.alignment = left
        cell.border = border_all

# Footer note
note_row = hr + 1 + len(qref_rows) + 1
ws.merge_cells(start_row=note_row, start_column=2, end_row=note_row, end_column=6)
c = ws.cell(row=note_row, column=2,
            value=("Note: All formulas reference CONFIG!$B$3 (ticker) and CONFIG!$B$4 (currency). "
                   "Never hardcode a ticker on any sheet. Bloomberg Excel Add-in is required for live data."))
c.font = font_note
c.fill = fill_note
c.alignment = left
ws.row_dimensions[note_row].height = 32


# ---------------- BETA sheet ----------------
ws = wb.create_sheet("BETA")
set_widths(ws, [4, 38, 22, 56])
ws.merge_cells("B1:D1")
ws["B1"] = "BETA — Raw, Adjusted, and R-squared"
ws["B1"].font = font_title
ws["B1"].fill = fill_title
ws["B1"].alignment = center
ws.row_dimensions[1].height = 24

ws["B3"] = "Ticker"
ws["B3"].font = font_label
ws["C3"] = "=CONFIG!$B$3"
ws["C3"].font = Font(name=ARIAL, size=10, color="008000")  # green = cross-sheet link

headers = ["Metric", "Value", "Bloomberg Field"]
for j, h in enumerate(headers, start=2):
    cell = ws.cell(row=5, column=j, value=h)
    cell.font = font_label
    cell.fill = fill_header
    cell.alignment = center
    cell.border = border_all

beta_rows = [
    ("Raw Beta (overridable, 2Y weekly vs local index)",
     '=BDP(CONFIG!$B$3,"BETA_RAW_OVERRIDABLE")', "BETA_RAW_OVERRIDABLE"),
    ("Adjusted Beta (Bloomberg adjusted)",
     '=BDP(CONFIG!$B$3,"BETA_ADJ_OVERRIDABLE")', "BETA_ADJ_OVERRIDABLE"),
    ("R-squared",
     '=BDP(CONFIG!$B$3,"R_SQR")', "R_SQR"),
]
for i, (label, formula, field) in enumerate(beta_rows, start=6):
    ws.cell(row=i, column=2, value=label).font = font_body
    fcell = ws.cell(row=i, column=3, value=formula)
    fcell.font = font_formula
    fcell.alignment = center
    fcell.number_format = "0.000"
    ws.cell(row=i, column=4, value=field).font = font_note
    for j in range(2, 5):
        ws.cell(row=i, column=j).border = border_all


# ---------------- DVD sheet ----------------
ws = wb.create_sheet("DVD")
set_widths(ws, [4, 16, 14, 14, 14, 12, 14, 16])
ws.merge_cells("B1:H1")
ws["B1"] = "DVD — Dividend History"
ws["B1"].font = font_title
ws["B1"].fill = fill_title
ws["B1"].alignment = center
ws.row_dimensions[1].height = 24

ws["B3"] = "Ticker"
ws["B3"].font = font_label
ws["C3"] = "=CONFIG!$B$3"
ws["C3"].font = Font(name=ARIAL, size=10, color="008000")

# BDS array formula — spills down/right starting at B6
note = ws.cell(row=4, column=2,
               value=("BDS returns the full dividend history. The array spills from B6. "
                      "Columns: Declared Date | Ex-Date | Record Date | Payable Date | Amount | Frequency | Type"))
note.font = font_note
note.alignment = left
ws.merge_cells("B4:H4")
ws.row_dimensions[4].height = 28

dvd_headers = ["Declared Date", "Ex-Date", "Record Date", "Payable Date", "Amount", "Frequency", "Type"]
for j, h in enumerate(dvd_headers, start=2):
    cell = ws.cell(row=5, column=j, value=h)
    cell.font = font_label
    cell.fill = fill_header
    cell.alignment = center
    cell.border = border_all

# The BDS spill formula — single cell, spills the whole table
ws["B6"] = '=BDS(CONFIG!$B$3,"DVD_HIST_ALL")'
ws["B6"].font = font_formula
ws["B6"].alignment = center


# ---------------- EE sheet ----------------
ws = wb.create_sheet("EE")
set_widths(ws, [4, 38, 22, 56])
ws.merge_cells("B1:D1")
ws["B1"] = "EE — Earnings Estimates (Forward Consensus)"
ws["B1"].font = font_title
ws["B1"].fill = fill_title
ws["B1"].alignment = center
ws.row_dimensions[1].height = 24

ws["B3"] = "Ticker"; ws["B3"].font = font_label
ws["C3"] = "=CONFIG!$B$3"; ws["C3"].font = Font(name=ARIAL, size=10, color="008000")
ws["B4"] = "Currency"; ws["B4"].font = font_label
ws["C4"] = "=CONFIG!$B$4"; ws["C4"].font = Font(name=ARIAL, size=10, color="008000")

for j, h in enumerate(["Metric", "Value", "Bloomberg Field"], start=2):
    cell = ws.cell(row=6, column=j, value=h)
    cell.font = font_label; cell.fill = fill_header
    cell.alignment = center; cell.border = border_all

ee_rows = [
    ("Forward EPS (next fiscal year, BEst)",
     '=BDP(CONFIG!$B$3,"BEST_EPS","BEST_FPERIOD_OVERRIDE","1FY","EQY_FUND_CRNCY",CONFIG!$B$4)',
     "BEST_EPS"),
    ("Long-Term Growth (LTG, %)",
     '=BDP(CONFIG!$B$3,"BEST_LTG_EPS")', "BEST_LTG_EPS"),
    ("Revenue Estimate (next fiscal year, BEst)",
     '=BDP(CONFIG!$B$3,"BEST_SALES","BEST_FPERIOD_OVERRIDE","1FY","EQY_FUND_CRNCY",CONFIG!$B$4)',
     "BEST_SALES"),
    ("EBITDA Estimate (next fiscal year, BEst)",
     '=BDP(CONFIG!$B$3,"BEST_EBITDA","BEST_FPERIOD_OVERRIDE","1FY","EQY_FUND_CRNCY",CONFIG!$B$4)',
     "BEST_EBITDA"),
    ("Consensus Rating (1=Buy, 5=Sell)",
     '=BDP(CONFIG!$B$3,"BEST_ANALYST_RATING")', "BEST_ANALYST_RATING"),
]
for i, (label, formula, field) in enumerate(ee_rows, start=7):
    ws.cell(row=i, column=2, value=label).font = font_body
    fcell = ws.cell(row=i, column=3, value=formula)
    fcell.font = font_formula; fcell.alignment = center
    ws.cell(row=i, column=4, value=field).font = font_note
    for j in range(2, 5):
        ws.cell(row=i, column=j).border = border_all
ws.cell(row=8, column=3).number_format = "0.0%"  # LTG as %


# ---------------- RV_Comps sheet ----------------
ws = wb.create_sheet("RV_Comps")
set_widths(ws, [4, 38, 22, 56])
ws.merge_cells("B1:D1")
ws["B1"] = "RV_Comps — Relative Valuation Multiples"
ws["B1"].font = font_title
ws["B1"].fill = fill_title
ws["B1"].alignment = center
ws.row_dimensions[1].height = 24

ws["B3"] = "Ticker"; ws["B3"].font = font_label
ws["C3"] = "=CONFIG!$B$3"; ws["C3"].font = Font(name=ARIAL, size=10, color="008000")

for j, h in enumerate(["Multiple", "Value", "Bloomberg Field"], start=2):
    cell = ws.cell(row=5, column=j, value=h)
    cell.font = font_label; cell.fill = fill_header
    cell.alignment = center; cell.border = border_all

rv_rows = [
    ("EV / EBITDA (current)",
     '=BDP(CONFIG!$B$3,"CURRENT_EV_BEST_EBITDA")', "CURRENT_EV_BEST_EBITDA"),
    ("P / E (BEst forward)",
     '=BDP(CONFIG!$B$3,"BEST_PE_RATIO")', "BEST_PE_RATIO"),
    ("EV / Sales (current)",
     '=BDP(CONFIG!$B$3,"EV_TO_T12M_SALES")', "EV_TO_T12M_SALES"),
    ("Price / Book",
     '=BDP(CONFIG!$B$3,"PX_TO_BOOK_RATIO")', "PX_TO_BOOK_RATIO"),
]
for i, (label, formula, field) in enumerate(rv_rows, start=6):
    ws.cell(row=i, column=2, value=label).font = font_body
    fcell = ws.cell(row=i, column=3, value=formula)
    fcell.font = font_formula; fcell.alignment = center
    fcell.number_format = "0.0\"x\""
    ws.cell(row=i, column=4, value=field).font = font_note
    for j in range(2, 5):
        ws.cell(row=i, column=j).border = border_all


# ---------------- ANR sheet ----------------
ws = wb.create_sheet("ANR")
set_widths(ws, [4, 38, 22, 56])
ws.merge_cells("B1:D1")
ws["B1"] = "ANR — Analyst Recommendations"
ws["B1"].font = font_title
ws["B1"].fill = fill_title
ws["B1"].alignment = center
ws.row_dimensions[1].height = 24

ws["B3"] = "Ticker"; ws["B3"].font = font_label
ws["C3"] = "=CONFIG!$B$3"; ws["C3"].font = Font(name=ARIAL, size=10, color="008000")
ws["B4"] = "Currency"; ws["B4"].font = font_label
ws["C4"] = "=CONFIG!$B$4"; ws["C4"].font = Font(name=ARIAL, size=10, color="008000")

for j, h in enumerate(["Metric", "Value", "Bloomberg Field"], start=2):
    cell = ws.cell(row=6, column=j, value=h)
    cell.font = font_label; cell.fill = fill_header
    cell.alignment = center; cell.border = border_all

anr_rows = [
    ("Consensus Rating (1=Buy ... 5=Sell)",
     '=BDP(CONFIG!$B$3,"BEST_ANALYST_RATING")', "BEST_ANALYST_RATING", "0.00"),
    ("Buy recommendations",
     '=BDP(CONFIG!$B$3,"TOT_BUY_REC")', "TOT_BUY_REC", "0"),
    ("Hold recommendations",
     '=BDP(CONFIG!$B$3,"TOT_HOLD_REC")', "TOT_HOLD_REC", "0"),
    ("Sell recommendations",
     '=BDP(CONFIG!$B$3,"TOT_SELL_REC")', "TOT_SELL_REC", "0"),
    ("12M Target Price (consensus)",
     '=BDP(CONFIG!$B$3,"BEST_TARGET_PRICE","EQY_FUND_CRNCY",CONFIG!$B$4)',
     "BEST_TARGET_PRICE", "#,##0.00"),
    ("Coverage (total analysts)",
     '=BDP(CONFIG!$B$3,"TOT_ANALYST_REC")', "TOT_ANALYST_REC", "0"),
]
for i, (label, formula, field, fmt) in enumerate(anr_rows, start=7):
    ws.cell(row=i, column=2, value=label).font = font_body
    fcell = ws.cell(row=i, column=3, value=formula)
    fcell.font = font_formula; fcell.alignment = center
    fcell.number_format = fmt
    ws.cell(row=i, column=4, value=field).font = font_note
    for j in range(2, 5):
        ws.cell(row=i, column=j).border = border_all


# ---------------- HP_Monthly sheet ----------------
ws = wb.create_sheet("HP_Monthly")
set_widths(ws, [4, 16, 16, 60])
ws.merge_cells("B1:D1")
ws["B1"] = "HP_Monthly — Historical Prices (10-Year, Monthly)"
ws["B1"].font = font_title
ws["B1"].fill = fill_title
ws["B1"].alignment = center
ws.row_dimensions[1].height = 24

ws["B3"] = "Ticker"; ws["B3"].font = font_label
ws["C3"] = "=CONFIG!$B$3"; ws["C3"].font = Font(name=ARIAL, size=10, color="008000")

note = ws.cell(row=4, column=2,
               value=("BDH returns a 2-column array (Date | PX_LAST) that spills from B6. "
                      "10-year monthly close, last day of month."))
note.font = font_note; note.alignment = left
ws.merge_cells("B4:D4"); ws.row_dimensions[4].height = 28

for j, h in enumerate(["Date", "PX_LAST"], start=2):
    cell = ws.cell(row=5, column=j, value=h)
    cell.font = font_label; cell.fill = fill_header
    cell.alignment = center; cell.border = border_all

ws["B6"] = '=BDH(CONFIG!$B$3,"PX_LAST",TODAY()-3653,TODAY(),"Per=cm","Days=A","Fill=P")'
ws["B6"].font = font_formula
ws["B6"].alignment = center


# ---------------- HP_Daily sheet ----------------
ws = wb.create_sheet("HP_Daily")
set_widths(ws, [4, 16, 16, 60])
ws.merge_cells("B1:D1")
ws["B1"] = "HP_Daily — Historical Prices (5-Year, Daily)"
ws["B1"].font = font_title
ws["B1"].fill = fill_title
ws["B1"].alignment = center
ws.row_dimensions[1].height = 24

ws["B3"] = "Ticker"; ws["B3"].font = font_label
ws["C3"] = "=CONFIG!$B$3"; ws["C3"].font = Font(name=ARIAL, size=10, color="008000")

note = ws.cell(row=4, column=2,
               value=("BDH returns a 2-column array (Date | PX_LAST) that spills from B6. "
                      "5-year daily close, weekdays only."))
note.font = font_note; note.alignment = left
ws.merge_cells("B4:D4"); ws.row_dimensions[4].height = 28

for j, h in enumerate(["Date", "PX_LAST"], start=2):
    cell = ws.cell(row=5, column=j, value=h)
    cell.font = font_label; cell.fill = fill_header
    cell.alignment = center; cell.border = border_all

ws["B6"] = '=BDH(CONFIG!$B$3,"PX_LAST",TODAY()-1826,TODAY(),"Per=cd","Days=W","Fill=P")'
ws["B6"].font = font_formula
ws["B6"].alignment = center


# Make CONFIG the active sheet on open
wb.active = wb.sheetnames.index("CONFIG")

out_path = r"C:\Users\MagnusThomassen\MFIP_Bloomberg_Export_Template_Master.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
