from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ARIAL = "Arial"
font_title = Font(name=ARIAL, size=14, bold=True)
font_label = Font(name=ARIAL, size=10, bold=True)
font_body = Font(name=ARIAL, size=10)
font_input = Font(name=ARIAL, size=11, bold=True)
font_header = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")
font_formula = Font(name=ARIAL, size=10)

fill_input = PatternFill("solid", start_color="F8CBAD")
fill_header = PatternFill("solid", start_color="305496")
thin = Side(border_style="thin", color="BFBFBF")
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center")
left_align = Alignment(horizontal="left", vertical="center")

wb = Workbook()

# ---------------- CONFIG ----------------
ws = wb.active
ws.title = "CONFIG"
ws.column_dimensions["A"].width = 32
ws.column_dimensions["B"].width = 24

ws["A1"] = "MFIP Bloomberg FX Export Template"
ws["A1"].font = font_title

ws["A3"] = "EXPORT_DATE:"
ws["A3"].font = font_label
ws["A3"].alignment = left_align

ws["B3"] = date.today().strftime("%Y-%m-%d")
ws["B3"].font = font_input
ws["B3"].fill = fill_input
ws["B3"].alignment = center
ws["B3"].border = border_all

ws["A6"] = "INSTRUCTIONS:"
ws["A6"].font = font_label

instructions = [
    (7,  "1. Run once per session (same session as company and indices exports)"),
    (8,  "2. Export date in B3 should match the date in your company file names"),
    (9,  "3. Wait 15 seconds for all BDH formulas to populate"),
    (10, "4. Select all (Ctrl+A), Paste Special > Values"),
    (11, "5. Save as: FX_<YYYY-MM-DD>.xlsx (e.g., FX_2026-05-09.xlsx)"),
    (13, "NOTE: These FX crosses are used for reconciling PDF financial statements"),
    (14, "(which are in local currency) with Bloomberg market data (which are in USD)."),
]
for r, txt in instructions:
    ws.cell(row=r, column=1, value=txt).font = font_body

# ---------------- FX sheets ----------------
FX_CROSSES = [
    ("NOKUSD", "NOKUSD Curncy", "Norwegian Krone to US Dollar — used for EQNR, DNB, TEL"),
    ("DKKUSD", "DKKUSD Curncy", "Danish Krone to US Dollar — used for NOVO B"),
    ("GBPUSD", "GBPUSD Curncy", "British Pound to US Dollar — used for CKN"),
]

for sheet_name, ticker, desc in FX_CROSSES:
    s = wb.create_sheet(sheet_name)
    s.column_dimensions["A"].width = 14
    s.column_dimensions["B"].width = 16

    h1 = s.cell(row=1, column=1, value="Date")
    h2 = s.cell(row=1, column=2, value="Rate")
    for h in (h1, h2):
        h.font = font_header
        h.fill = fill_header
        h.alignment = center
        h.border = border_all

    f = s.cell(row=2, column=1,
               value=f'=BDH("{ticker}","PX_LAST","20160101","20260508","MONTHLY")')
    f.font = font_formula
    f.alignment = center

wb.active = wb.sheetnames.index("CONFIG")

out = r"C:\Users\MagnusThomassen\MFIP_Bloomberg_Export_Template_FX.xlsx"
wb.save(out)
print(f"Saved: {out}")
