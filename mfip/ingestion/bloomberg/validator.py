"""Validate a saved Bloomberg company workbook against the v1 export contract.

Contract source: 07_BLOOMBERG_EXPORT_TEMPLATE.docx (REQUIRED SHEETS) and
decisions.md 2026-05-09 "Master Bloomberg template structure locked in".

A saved company workbook has 8 sheets:
    CONFIG, BETA, DVD, EE, ANR, HP_Monthly, HP_Daily, RV_Comps

CONFIG carries the ticker (B3) and currency (B4) that drove the export.
The other seven sheets carry data; CONFIG is metadata, skipped by the
Phase 3 ingestion parser but validated here for self-documentation.

This script is a one-shot pre-Phase-3 validator. It proves the contract
holds end-to-end before the full ingestion layer is built. Two modes:

    --file <path>       Validate one workbook explicitly.
    --all               Walk bloomberg_archive\\, pick latest date per
                        company folder, validate each. Loud-flag any
                        universe member that has no folder at all.

Three result tiers, mapped to the Security Council alert system:

    PASS        Contract holds.
    ADVISORY    Non-fatal drift (e.g. RV_Comps column missing). The
                Comps Agent handles missing multiples downstream.
    FAIL        Contract broken (wrong sheet names, missing ticker,
                empty data sheets). File should be quarantined.

Usage from venv:
    python -m mfip.ingestion.bloomberg.validator --all
    python scripts\\ingestion\\validate_bloomberg_workbook.py --all   # legacy shell, unchanged behaviour

The latter remains as a thin CLI wrapper for compatibility with existing
documentation and any operator muscle memory; the former is the canonical
form going forward.
"""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass, field
from enum import Enum
from pathlib import Path

from openpyxl import load_workbook

from mfip.ingestion.bloomberg.archive_lookup import (
    KNOWN_ABSENT,
    UNIVERSE,
    find_all_latest,
    find_latest_workbook,
)


# ---------------------------------------------------------------------------
# Contract definitions
# ---------------------------------------------------------------------------

# v1 universe. Folder/file prefix matches Bloomberg-style ticker-with-suffix.
# CONFIG!B4 is the *display* currency, standardised to USD across all
# exports per decisions.md 2026-05-08 "Bloomberg export: USD standardisation
# and FX source". The local listing currency is metadata that lives outside
# the export (it's derived from the ticker exchange suffix during Phase 3
# ingestion, not from the workbook). The BETA sheet is the only sheet that
# stays in local currency, and it doesn't carry a CONFIG!B4 value of its own.
EXPECTED_CONFIG_CURRENCY: str = "USD"

# UNIVERSE and KNOWN_ABSENT are defined in archive_lookup.py (archive-shape
# concerns) and re-exported via the import above so callers that historically
# imported them from validator continue to work.

REQUIRED_SHEETS: list[str] = [
    "CONFIG", "BETA", "DVD", "EE", "ANR", "HP_Monthly", "HP_Daily", "RV_Comps",
]

DATA_SHEETS: list[str] = [s for s in REQUIRED_SHEETS if s != "CONFIG"]

# RV_Comps column sets. Per the 2026-05-09 "Master Bloomberg template
# structure locked in" decision, this is a terminal-side contract — Phase 3
# matches by header name (not position) and surfaces drift as Advisory.
#
# Names match the actual Bloomberg-native column headers as they land in
# the export (verified against EQNR_NO_2026-05-08.xlsx). decisions.md
# 2026-05-09 originally listed simplified names ("P/Book" rather than
# "Px/Book", "Rev 1Y growth" rather than "Rev - 1 Yr Gr:Y"); the export is
# authoritative, decisions.md will be corrected.
#
# Two-tier set:
#   BASELINE — must be present today. Missing => Advisory (RV_COLUMN_DRIFT).
#   PLANNED  — known-pending columns. Scheduled for addition at the next
#              lab session. Missing => Advisory with distinct code so the
#              gap stays visible without being noisy. Once added, promote
#              from PLANNED to BASELINE in this file.
RV_BASELINE_COLUMNS: list[str] = [
    "Ticker", "Name", "Mkt Cap (USD)", "Last Px (USD)",
    "P/E", "ROE", "Dvd 12M Yld",
    "Rev - 1 Yr Gr:Y", "EPS - 1 Yr Gr:Y",
]

# Pending: to be added to the terminal-side RV layout at the next Kingston
# lab session. Missing => RV_PLANNED_COLUMN_PENDING (Advisory, informational).
RV_PLANNED_COLUMNS: list[str] = [
    "EV/EBITDA", "Px/Book", "EV/Sales",
]

FILENAME_RE = re.compile(r"^(?P<prefix>[A-Z0-9_]+)_(?P<date>\d{4}-\d{2}-\d{2})\.xlsx$")


# ---------------------------------------------------------------------------
# Result types
# ---------------------------------------------------------------------------

class Tier(Enum):
    PASS = "PASS"
    ADVISORY = "ADVISORY"
    FAIL = "FAIL"


@dataclass
class Finding:
    tier: Tier
    code: str
    message: str

    def __str__(self) -> str:
        return f"  [{self.tier.value}] {self.code}: {self.message}"


@dataclass
class ValidationReport:
    file: Path
    findings: list[Finding] = field(default_factory=list)

    def add(self, tier: Tier, code: str, message: str) -> None:
        self.findings.append(Finding(tier, code, message))

    @property
    def tier(self) -> Tier:
        """Worst tier across all findings. Empty report = PASS."""
        if any(f.tier == Tier.FAIL for f in self.findings):
            return Tier.FAIL
        if any(f.tier == Tier.ADVISORY for f in self.findings):
            return Tier.ADVISORY
        return Tier.PASS

    def render(self) -> str:
        lines = [f"{self.tier.value}  {self.file}"]
        for f in self.findings:
            lines.append(str(f))
        return "\n".join(lines)


# ---------------------------------------------------------------------------
# Validators (one concern per function, composed)
# ---------------------------------------------------------------------------

def _validate_filename(path: Path, report: ValidationReport) -> tuple[str | None, str | None]:
    """Returns (prefix, date) parsed from the filename, or (None, None) on FAIL."""
    m = FILENAME_RE.match(path.name)
    if not m:
        report.add(
            Tier.FAIL, "FILENAME",
            f"filename {path.name!r} does not match <PREFIX>_<YYYY-MM-DD>.xlsx",
        )
        return None, None
    prefix = m.group("prefix")
    date = m.group("date")

    # Filename prefix should match the parent folder structure:
    # bloomberg_archive\<PREFIX>\<DATE>\<PREFIX>_<DATE>.xlsx
    if path.parent.name != date:
        report.add(
            Tier.ADVISORY, "PATH_DATE_MISMATCH",
            f"filename date {date!r} != parent folder {path.parent.name!r}",
        )
    expected_company_folder = path.parent.parent.name
    if prefix != expected_company_folder:
        report.add(
            Tier.ADVISORY, "PATH_PREFIX_MISMATCH",
            f"filename prefix {prefix!r} != company folder {expected_company_folder!r}",
        )
    return prefix, date


def _validate_sheet_names(wb, report: ValidationReport) -> None:
    actual = set(wb.sheetnames)
    expected = set(REQUIRED_SHEETS)
    missing = expected - actual
    extra = actual - expected
    if missing:
        report.add(
            Tier.FAIL, "MISSING_SHEETS",
            f"required sheets absent: {sorted(missing)}",
        )
    if extra:
        # Extra sheets are an Advisory, not a Fail — the parser ignores them,
        # but the user should know the workbook has more than the contract
        # specifies (could be a save-as-values mistake).
        report.add(
            Tier.ADVISORY, "EXTRA_SHEETS",
            f"sheets present beyond the contract: {sorted(extra)}",
        )


def _validate_config(wb, prefix: str | None, report: ValidationReport) -> None:
    """Validate CONFIG!B3 (ticker) and CONFIG!B4 (currency)."""
    if "CONFIG" not in wb.sheetnames:
        return  # already reported by sheet-name check
    ws = wb["CONFIG"]
    config_ticker = ws["B3"].value
    config_currency = ws["B4"].value

    if not config_ticker or not isinstance(config_ticker, str):
        report.add(
            Tier.FAIL, "CONFIG_TICKER_MISSING",
            f"CONFIG!B3 is empty or non-string: {config_ticker!r}",
        )
    elif prefix is not None:
        # CONFIG!B3 holds Bloomberg-format tickers like "EQNR NO Equity".
        # We require the file-prefix-derived ticker (e.g. "EQNR_NO" -> "EQNR NO")
        # to appear at the start.
        normalised = config_ticker.strip().upper()
        expected_start = prefix.replace("_", " ").upper()
        if not normalised.startswith(expected_start):
            report.add(
                Tier.FAIL, "CONFIG_TICKER_MISMATCH",
                f"CONFIG!B3 {config_ticker!r} does not start with {expected_start!r}",
            )

    if not config_currency or not isinstance(config_currency, str):
        report.add(
            Tier.ADVISORY, "CONFIG_CURRENCY_MISSING",
            f"CONFIG!B4 is empty or non-string: {config_currency!r}",
        )
    elif config_currency.strip().upper() != EXPECTED_CONFIG_CURRENCY.upper():
        # All Bloomberg exports are USD-standardised per the 2026-05-08
        # decision. A non-USD value here means the lab session diverged
        # from the standardisation rule.
        report.add(
            Tier.ADVISORY, "CONFIG_CURRENCY_NOT_STANDARDISED",
            f"CONFIG!B4 {config_currency!r} != standardised {EXPECTED_CONFIG_CURRENCY!r} "
            "(see decisions.md 2026-05-08 USD standardisation)",
        )





def _validate_data_sheets_nonempty(wb, report: ValidationReport) -> None:
    """Each data sheet must have at least one row of numeric/date values.

    Catches a paste-as-values failure where formulas were stripped but the
    underlying data didn't paste through, leaving header rows only.

    Layout-agnostic: the master template has titles on row 1, headers on
    row 3, data from row 4; RV_Comps (manual paste) has headers on row 1,
    data from row 2. Rather than encode either layout, we scan the whole
    used range for a row that contains at least one non-string value
    (numeric, date, bool, etc.). Header/title rows are typically all
    strings; data rows are not.
    """
    for name in DATA_SHEETS:
        if name not in wb.sheetnames:
            continue  # already reported
        ws = wb[name]
        scan_max = min(ws.max_row or 0, 50)  # 50 is plenty to find one data row
        max_col = min(ws.max_column or 0, 12)
        if scan_max == 0 or max_col == 0:
            report.add(
                Tier.FAIL, "EMPTY_DATA_SHEET",
                f"sheet {name!r} is empty",
            )
            continue
        found_data = False
        for row in ws.iter_rows(min_row=1, max_row=scan_max,
                                max_col=max_col, values_only=True):
            for v in row:
                # A non-string, non-None, non-empty value is data.
                if v is None:
                    continue
                if isinstance(v, str) and v.strip() == "":
                    continue
                if not isinstance(v, str):
                    found_data = True
                    break
            if found_data:
                break
        if not found_data:
            report.add(
                Tier.FAIL, "EMPTY_DATA_SHEET",
                f"sheet {name!r} has no numeric/date values "
                f"(strings only — possible paste-as-values failure)",
            )


def _validate_rv_columns(wb, report: ValidationReport) -> None:
    """RV_Comps column-set drift surfaces as Advisory (per 2026-05-09 decision).

    Match by header name, not position. The header row varies in the export:
    EQNR_NO_2026-05-08 has metadata on rows 1-2 ('EQY_FUND_CRNCY', 'USD')
    and the actual column headers on row 3. Detect the header row as the
    first row in the first 6 rows that contains BOTH 'Ticker' AND 'Name'
    (insufficient to match just one — 'Ticker' alone appears in the
    'None (9 securities)' summary row in some exports).

    Two-tier check: BASELINE columns missing => Advisory (real drift);
    PLANNED columns missing => Advisory with distinct code (informational,
    "scheduled to be added next lab session").
    """
    if "RV_Comps" not in wb.sheetnames:
        return
    ws = wb["RV_Comps"]

    header_row_values: list[str] = []
    for row in ws.iter_rows(min_row=1, max_row=6,
                            max_col=min(ws.max_column or 1, 30),
                            values_only=True):
        cells = [str(c).strip() if c is not None else "" for c in row]
        cells_lower = {c.lower() for c in cells if c}
        if "ticker" in cells_lower and "name" in cells_lower:
            header_row_values = cells
            break

    if not header_row_values:
        report.add(
            Tier.ADVISORY, "RV_HEADERS_NOT_FOUND",
            "RV_Comps: could not locate header row (no row in first 6 rows "
            "contains both 'Ticker' and 'Name')",
        )
        return

    headers_lower = {h.lower() for h in header_row_values if h}

    missing_baseline = [c for c in RV_BASELINE_COLUMNS if c.lower() not in headers_lower]
    if missing_baseline:
        report.add(
            Tier.ADVISORY, "RV_COLUMN_DRIFT",
            f"RV_Comps missing baseline columns: {missing_baseline}. "
            "Comps Agent will handle per sector-appropriate selection.",
        )

    missing_planned = [c for c in RV_PLANNED_COLUMNS if c.lower() not in headers_lower]
    if missing_planned:
        report.add(
            Tier.ADVISORY, "RV_PLANNED_COLUMN_PENDING",
            f"RV_Comps missing planned columns: {missing_planned}. "
            "Scheduled for addition at next lab session — when added, "
            "promote from RV_PLANNED_COLUMNS to RV_BASELINE_COLUMNS.",
        )


# ---------------------------------------------------------------------------
# Top-level entry
# ---------------------------------------------------------------------------

def validate_workbook(path: Path) -> ValidationReport:
    report = ValidationReport(file=path)

    if not path.is_file():
        report.add(Tier.FAIL, "FILE_NOT_FOUND", f"{path} does not exist")
        return report
    if path.suffix.lower() != ".xlsx":
        report.add(Tier.FAIL, "WRONG_EXTENSION", f"expected .xlsx, got {path.suffix!r}")
        return report

    prefix, _date = _validate_filename(path, report)

    try:
        # data_only=True so we read paste-as-values cached values, not formulas.
        # read_only=True keeps memory bounded for the larger workbooks.
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception as e:  # noqa: BLE001 - want broad catch for corrupt files
        report.add(Tier.FAIL, "OPEN_FAILED", f"openpyxl could not open: {e}")
        return report

    try:
        _validate_sheet_names(wb, report)
        _validate_config(wb, prefix, report)
        _validate_data_sheets_nonempty(wb, report)
        _validate_rv_columns(wb, report)
    finally:
        wb.close()

    return report


def validate_all(archive_root: Path) -> list[ValidationReport]:
    """Walk bloomberg_archive\\, validate latest workbook per universe member.

    Uses the policy-free find_all_latest walker from archive_lookup.py
    and applies known-absent policy here: folder-missing-and-known-absent
    surfaces as ADVISORY; folder-missing-and-unknown is FAIL;
    folder-present-but-no-dated-export is FAIL.
    """
    reports: list[ValidationReport] = []
    latest_by_prefix = find_all_latest(archive_root)
    # Iterate in sorted order so output is deterministic across Python
    # processes. UNIVERSE is a set, so the pre-refactor iteration order
    # was randomized by Python's per-process hash seed; the visible
    # report ordering changed every run. PR #62a introduces sorted
    # iteration here — strictly an observability improvement, no impact
    # on findings, counts, or exit codes (the smoke test verifies this).
    for prefix in sorted(latest_by_prefix):
        latest = latest_by_prefix[prefix]
        if latest is not None:
            reports.append(validate_workbook(latest))
            continue
        # latest is None — disambiguate folder-missing vs no-dated-export.
        company_dir = archive_root / prefix
        r = ValidationReport(file=company_dir)
        if not company_dir.is_dir():
            if prefix in KNOWN_ABSENT:
                r.add(
                    Tier.ADVISORY, "UNIVERSE_MEMBER_ABSENT",
                    f"{prefix} folder not present in archive (known absent: "
                    "see decisions.md / next lab session)",
                )
            else:
                r.add(
                    Tier.FAIL, "UNIVERSE_MEMBER_ABSENT",
                    f"{prefix} folder not present in archive",
                )
        else:
            r.add(
                Tier.FAIL, "NO_DATED_EXPORT",
                f"{prefix}: no <YYYY-MM-DD>\\*.xlsx found",
            )
        reports.append(r)
    return reports


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    g = p.add_mutually_exclusive_group(required=True)
    g.add_argument("--file", type=Path, help="Validate a single workbook.")
    g.add_argument("--all", action="store_true",
                   help="Validate latest workbook per universe member.")
    p.add_argument(
        "--archive-root", type=Path,
        default=Path(r"C:\MFIP\bloomberg_archive"),
        help="Root of bloomberg_archive (only used with --all).",
    )
    args = p.parse_args()

    if args.file:
        reports = [validate_workbook(args.file)]
    else:
        reports = validate_all(args.archive_root)

    # Print human-readable report.
    print("=" * 72)
    print("MFIP Bloomberg workbook validator — saved-file contract")
    print("=" * 72)
    for r in reports:
        print()
        print(r.render())

    # Summary line + exit code.
    counts = {t: 0 for t in Tier}
    for r in reports:
        counts[r.tier] += 1
    print()
    print("-" * 72)
    print(
        f"Summary: PASS={counts[Tier.PASS]}  "
        f"ADVISORY={counts[Tier.ADVISORY]}  "
        f"FAIL={counts[Tier.FAIL]}"
    )

    # Exit non-zero on any FAIL; ADVISORY is informational.
    return 1 if counts[Tier.FAIL] else 0


if __name__ == "__main__":
    sys.exit(main())
