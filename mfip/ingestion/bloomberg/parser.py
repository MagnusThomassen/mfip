"""Bloomberg workbook parser — public entrypoint.

The canonical API is:

    parse_workbook(path: Path, *, validation_policy=ValidationPolicy.STRICT) -> ParsedCompanyData

The parser is a pure function: it does not write to decision_log,
does not send alerts, does not modify the input file. It emits
stdlib `logging` calls at INFO/WARNING/ERROR; callers route them.

Per-sheet extraction is delegated to private helpers; PR #62b ships
the wiring with CONFIG live and other sheets stubbed. PRs #63-#65
fill in the sheet-extraction bodies.
"""

import argparse
import logging
import sys
from datetime import date, datetime, timezone
from enum import Enum
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from mfip.ingestion.bloomberg.exceptions import (
    BloombergParserError,
    WorkbookExtractionError,
    WorkbookValidationError,
)
from mfip.ingestion.bloomberg.models import (
    AnalystRecommendationsSheet,
    BetaSheet,
    DividendSheet,
    EarningsEstimatesSheet,
    ParsedCompanyData,
    ParsedFXData,
    ParsedIndicesData,
    PriceHistorySheet,
    RVCompsSheet,
)
from mfip.ingestion.bloomberg.validator import (
    FILENAME_RE,
    Tier,
    validate_workbook,
)

logger = logging.getLogger(__name__)


class ValidationPolicy(Enum):
    """Controls parser behaviour when the validator returns a FAIL.

    STRICT (default): FAIL raises WorkbookValidationError. ADVISORY
        logged and carried forward into ParsedCompanyData.advisories.
        PASS proceeds silently. The v2-watchdog-friendly default.

    PERMISSIVE: FAIL logged as WARNING, parser attempts extraction
        anyway. For development / debugging only. Not used by agents.

    (REPORT_ONLY was specified in the original design doc but dropped
    in PR #62b — it duplicated PERMISSIVE's behaviour with no
    additional capability. Reintroduce when a consumer demonstrably
    needs it.)
    """

    STRICT = "strict"
    PERMISSIVE = "permissive"


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def parse_workbook(
    path: Path,
    *,
    validation_policy: ValidationPolicy = ValidationPolicy.STRICT,
) -> ParsedCompanyData:
    """Parse a Bloomberg company workbook into ParsedCompanyData.

    Per-sheet extraction is stubbed in PR #62b — every sub-model
    returns its default-empty form. CONFIG extraction (ticker,
    currency) is functional. validation_report and advisories are
    populated.

    Args:
        path: Path to the .xlsx workbook.
        validation_policy: How to handle a validator FAIL result.

    Returns:
        ParsedCompanyData with provenance, CONFIG fields, and
        validation_report populated; all per-sheet sub-models
        empty pending #63 onward.

    Raises:
        WorkbookValidationError: validator returned FAIL and policy
            is STRICT. NB: this catches the nonexistent-path case too
            (validator reports FILE_NOT_FOUND).
        WorkbookExtractionError: validator passed but openpyxl could
            not open the file, or CONFIG cells are malformed.
    """
    logger.info("Parsing workbook: %s", path)

    # 1. Validate first. The parser trusts the validator's contract
    #    check; it does not re-validate sheet presence or RV column
    #    shapes itself.
    report = validate_workbook(path)
    advisories = [f for f in report.findings if f.tier == Tier.ADVISORY]

    if report.tier == Tier.FAIL:
        if validation_policy == ValidationPolicy.STRICT:
            raise WorkbookValidationError(report)
        else:  # PERMISSIVE
            logger.warning(
                "Validator FAIL but PERMISSIVE policy — proceeding (%d findings)",
                len(report.findings),
            )
    elif report.tier == Tier.ADVISORY:
        logger.info("Validator ADVISORY: %d findings", len(advisories))

    # 2. Open the workbook. data_only=True reads cached values
    #    (paste-as-values output); read_only=True is faster.
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        raise WorkbookExtractionError("(workbook open)", str(exc)) from exc

    try:
        # 3. Extract CONFIG (ticker + currency). Other sheets are
        #    stubbed — they return default-empty sub-models pending
        #    PR #63 onward.
        ticker, ticker_short, config_currency = _extract_config(wb, path)

        result = ParsedCompanyData(
            source_path=path,
            parsed_at=datetime.now(timezone.utc),
            validation_report=report,
            advisories=advisories,
            ticker=ticker,
            ticker_short=ticker_short,
            config_currency=config_currency,
            listing_currency="",   # populated in PR #65
            is_gbp_pence=False,    # populated in PR #65
            beta=_extract_beta(wb),
            dvd=_extract_dvd(wb),
            ee=_extract_ee(wb),
            anr=_extract_anr(wb),
            hp_monthly=_extract_hp_monthly(wb),
            hp_daily=_extract_hp_daily(wb),
            rv_comps=_extract_rv_comps(wb),
        )
    finally:
        wb.close()

    logger.info("Parsed: ticker=%s advisories=%d", ticker_short, len(advisories))
    return result


def parse_indices_workbook(path: Path) -> ParsedIndicesData:
    """Parse the Bloomberg Indices workbook (4 index series).

    Skeleton in PR #62b — returns empty ParsedIndicesData. Full
    extraction lands in PR #66 (which will also add a
    validation_policy parameter once indices validation lands).
    """
    logger.info("Parsing indices workbook: %s", path)
    return ParsedIndicesData(
        source_path=path,
        parsed_at=datetime.now(timezone.utc),
    )


def parse_fx_workbook(path: Path) -> ParsedFXData:
    """Parse the Bloomberg FX workbook (3 FX cross series).

    Skeleton in PR #62b — returns empty ParsedFXData. Full extraction
    lands in PR #66 (which will also add a validation_policy parameter
    once FX validation lands).
    """
    logger.info("Parsing FX workbook: %s", path)
    return ParsedFXData(
        source_path=path,
        parsed_at=datetime.now(timezone.utc),
    )


# ---------------------------------------------------------------------------
# Private extraction helpers (CONFIG functional; others stubbed)
# ---------------------------------------------------------------------------

def _extract_config(wb, path: Path) -> tuple[str, str, str]:
    """Read CONFIG!B3 (ticker) and CONFIG!B4 (currency).

    Returns (ticker, ticker_short, config_currency).
    ticker_short prefers the filename prefix (validator's canonical
    contract via FILENAME_RE), falling back to ticker-string token
    splitting if the filename doesn't match the saved-file naming
    convention.
    """
    try:
        config = wb["CONFIG"]
    except KeyError as exc:
        raise WorkbookExtractionError("CONFIG", "sheet missing") from exc

    ticker = config["B3"].value
    currency = config["B4"].value

    if not isinstance(ticker, str) or not ticker.strip():
        raise WorkbookExtractionError(
            "CONFIG",
            f"B3 (ticker) malformed: {ticker!r}",
        )

    ticker_short = _derive_ticker_short(path, ticker)

    # Currency may be None / empty on the template (master workbook
    # before lab fill-in) or non-USD on a saved file (validator surfaces
    # this as ADVISORY, not FAIL — parser must not promote).
    config_currency = currency.strip() if isinstance(currency, str) else ""

    return ticker, ticker_short, config_currency


def _derive_ticker_short(path: Path, ticker_str: str) -> str:
    """Derive ticker_short. Filename is authoritative; ticker is fallback.

    The validator's FILENAME_RE matches the saved-file naming convention
    (e.g. "EQNR_NO_2026-05-08.xlsx") and exposes the prefix as a named
    group. When the filename matches, use that — it's the contract.

    Falls back to splitting the ticker string on whitespace and joining
    the first two tokens with underscore — works for the v1 universe
    but is cosmetically dependent on Bloomberg's ticker conventions.
    """
    match = FILENAME_RE.match(path.name)
    if match:
        return match.group("prefix")

    tokens = ticker_str.strip().split()
    if len(tokens) >= 2:
        return f"{tokens[0]}_{tokens[1]}"

    raise WorkbookExtractionError(
        "CONFIG",
        f"cannot derive ticker_short from filename={path.name!r} "
        f"or ticker={ticker_str!r}",
    )


# ---------------------------------------------------------------------------
# Scalar-cell helper (used by BETA, EE; reused by ANR scalars in #65).
# ---------------------------------------------------------------------------

# Bloomberg / Excel surfaces several variants of "no usable value":
#   #N/A, #N/A Requesting Data..., #N/A N/A, #N/A Invalid Security,
#   #N/A Field Not Applicable, #VALUE!, #NAME?, #REF!, #DIV/0!, #NUM!, #NULL!
# The enumerated set is documentation; the runtime check uses the broader
# "#"-prefix guard so future variants don't require a code change. The two
# together are belt-and-braces: the set lets readers see the known taxonomy,
# the prefix guard catches everything Excel/Bloomberg can throw.
_NA_VALUES = frozenset({
    "#N/A",
    "#N/A Requesting Data...",
    "#N/A N/A",
    "#N/A Invalid Security",
    "#N/A Field Not Applicable",
    "#VALUE!",
    "#NAME?",
    "#REF!",
    "#DIV/0!",
    "#NUM!",
    "#NULL!",
})


def _read_scalar(
    ws: Worksheet,
    coord: str,
    *,
    na_cells: list[str],
) -> float | None:
    """Read a single cell as float, recording NA-likes in `na_cells`.

    Returns None for any Excel/Bloomberg error string (any value starting
    with '#' after stripping), for None, or for empty string. Mutates
    `na_cells` in place: appends f"{ws.title}!{coord}" whenever None is
    returned. The "SHEET!COORD" format is intentional — when a caller
    needs to debug an NA value, they open the workbook and look at that
    cell. Field codes would require a lookup hop; model paths couple
    log content to the Pydantic API.

    Numeric strings ("1,234.5") are parsed permissively because Bloomberg
    occasionally emits comma-grouped values when locale settings drift.

    Raises:
        WorkbookExtractionError: cell holds a non-numeric, non-NA string.
            A value of 0.0 is a real number; None is "Bloomberg didn't
            return one"; an unexpected string like "approx 1.2" is a
            workbook corruption that should fail loud rather than be
            silently coerced to None.
    """
    raw: Any = ws[coord].value

    if raw is None or raw == "":
        na_cells.append(f"{ws.title}!{coord}")
        return None

    if isinstance(raw, str):
        stripped = raw.strip()
        # Any '#'-prefixed string is an Excel/Bloomberg error sentinel
        # (#N/A, #VALUE!, #NAME?, #REF!, #DIV/0!, #NUM!, #NULL!, and the
        # open-ended #N/A-family Bloomberg variants). Prefix guard is
        # correct here specifically because '#' is a closed lexical
        # convention — no real value can start with it in a paste-as-
        # values cell.
        if stripped in _NA_VALUES or stripped.startswith("#"):
            na_cells.append(f"{ws.title}!{coord}")
            return None
        try:
            return float(stripped.replace(",", ""))
        except ValueError as exc:
            raise WorkbookExtractionError(
                ws.title,
                f"non-numeric, non-NA value at {ws.title}!{coord}: {raw!r}",
            ) from exc

    if isinstance(raw, (int, float)):
        return float(raw)

    raise WorkbookExtractionError(
        ws.title,
        f"unexpected cell type at {ws.title}!{coord}: {type(raw).__name__} {raw!r}",
    )


# ---------------------------------------------------------------------------
# Per-sheet extraction helpers
# ---------------------------------------------------------------------------

def _extract_beta(wb) -> BetaSheet:
    """Extract BETA sheet scalars (PR #64).

    Coordinate map (header row 5):
        BETA!C6  raw_beta       (BETA_RAW_OVERRIDABLE)
        BETA!C7  adjusted_beta  (BETA_ADJ_OVERRIDABLE)
        BETA!C8  r_squared      (COEF_DETER_R_SQUARED)

    #N/A is expected for thin-coverage names on BETA!C8.
    """
    ws = wb["BETA"]
    na_cells: list[str] = []
    raw_beta = _read_scalar(ws, "C6", na_cells=na_cells)
    adjusted_beta = _read_scalar(ws, "C7", na_cells=na_cells)
    r_squared = _read_scalar(ws, "C8", na_cells=na_cells)
    return BetaSheet(
        raw_beta=raw_beta,
        adjusted_beta=adjusted_beta,
        r_squared=r_squared,
        na_cells=sorted(na_cells),
    )


def _extract_dvd(wb) -> DividendSheet:
    """STUB — populated in PR #65."""
    return DividendSheet()


def _extract_ee(wb) -> EarningsEstimatesSheet:
    """Extract EE sheet scalars (PR #64).

    Coordinate map (header row 6):
        EE!C7   best_eps             (BEST_EPS)
        EE!C8   best_ltg_eeps        (BEST_LTG_EEPS) — decimal, not %
        EE!C9   best_sales           (BEST_SALES)
        EE!C10  best_ebitda          (BEST_EBITDA)
        EE!C11  best_analyst_rating  (BEST_ANALYST_RATING)

    Unit note: the template formula for C8 is `=_xll.BDP(...)%`, so the
    paste-as-values cell carries a decimal (0.05 = 5%). Extraction is
    faithful to the cell; normalisation happens at the consumer layer.

    BEST_LTG_EEPS expected #N/A for DNB (no LTG consensus on banks)
    and CKN (small-cap thin coverage). BEST_ANALYST_RATING is extracted
    for forensic completeness but not consumed downstream — the real
    signal lives in ANR's raw broker counts (decisions.md 2026-05-08).
    """
    ws = wb["EE"]
    na_cells: list[str] = []
    best_eps = _read_scalar(ws, "C7", na_cells=na_cells)
    best_ltg_eeps = _read_scalar(ws, "C8", na_cells=na_cells)
    best_sales = _read_scalar(ws, "C9", na_cells=na_cells)
    best_ebitda = _read_scalar(ws, "C10", na_cells=na_cells)
    best_analyst_rating = _read_scalar(ws, "C11", na_cells=na_cells)
    return EarningsEstimatesSheet(
        best_eps=best_eps,
        best_ltg_eeps=best_ltg_eeps,
        best_sales=best_sales,
        best_ebitda=best_ebitda,
        best_analyst_rating=best_analyst_rating,
        na_cells=sorted(na_cells),
    )


def _extract_anr(wb) -> AnalystRecommendationsSheet:
    """STUB — populated in PR #65."""
    return AnalystRecommendationsSheet()


def _read_price_series(
    ws: Worksheet,
    *,
    start_row: int = 6,
    date_col: str = "B",
    value_col: str = "C",
    na_cells: list[str],
) -> pd.Series:
    """Read a Date | Value array spill into a pandas.Series (PR #65).

    Iterates from `start_row` and terminates at the first row where
    both `date_col` and `value_col` are None. Openpyxl's `max_row` is
    NOT a reliable signal: templates are pre-sized for the eventual
    spill (HP_Daily template has max_row=1310 with only B6 populated),
    so we look at the data instead.

    Row-level handling:
        * Date col `#`-prefixed string  -> end-of-spill marker. Record
          coord in na_cells and break. (Bloomberg array formulas emit
          their error sentinel into the first cell of the spill region
          when the spill never materialised; the `#` prefix is the same
          closed lexical convention as in _read_scalar.)
        * Date col is None, value present -> price-without-date is
          unusable. Skip the row entirely, record the date coord.
        * Date col not a datetime/date  -> raises WorkbookExtractionError.
        * Value col None or `#`-prefixed -> store NaN in the series,
          record the value coord.
        * Value col numeric string  -> permissive parse (matches
          _read_scalar's "1,234.5" path).
        * Value col unexpected type  -> raises WorkbookExtractionError.

    Returns:
        pandas.Series indexed by DatetimeIndex(name="Date"),
        name="PX_LAST", dtype="float64". Empty Series with the right
        shape if the spill was a single `#NAME?` row (template case).
    """
    dates: list = []
    prices: list[float] = []
    row = start_row
    while True:
        raw_date = ws[f"{date_col}{row}"].value
        raw_value = ws[f"{value_col}{row}"].value

        if raw_date is None and raw_value is None:
            break  # End of spill.

        if isinstance(raw_date, str) and raw_date.startswith("#"):
            # Excel/Bloomberg error sentinel in the date column means
            # the spill never materialised (template case). Treat as
            # end-of-spill rather than a date-typed row.
            na_cells.append(f"{ws.title}!{date_col}{row}")
            break

        if raw_date is None:
            # Price without a date is unusable; skip the row.
            na_cells.append(f"{ws.title}!{date_col}{row}")
            row += 1
            continue

        if not isinstance(raw_date, (datetime, date)):
            raise WorkbookExtractionError(
                ws.title,
                f"non-date value in date column at "
                f"{ws.title}!{date_col}{row}: {raw_date!r}",
            )

        # Date is good; handle the value.
        if raw_value is None or (
            isinstance(raw_value, str) and raw_value.startswith("#")
        ):
            na_cells.append(f"{ws.title}!{value_col}{row}")
            dates.append(raw_date)
            prices.append(float("nan"))
        elif isinstance(raw_value, (int, float)):
            dates.append(raw_date)
            prices.append(float(raw_value))
        elif isinstance(raw_value, str):
            try:
                parsed = float(raw_value.replace(",", "").strip())
            except ValueError as exc:
                raise WorkbookExtractionError(
                    ws.title,
                    f"non-numeric, non-NA value in price column at "
                    f"{ws.title}!{value_col}{row}: {raw_value!r}",
                ) from exc
            dates.append(raw_date)
            prices.append(parsed)
        else:
            raise WorkbookExtractionError(
                ws.title,
                f"unexpected type in price column at "
                f"{ws.title}!{value_col}{row}: "
                f"{type(raw_value).__name__} {raw_value!r}",
            )
        row += 1

    return pd.Series(
        prices,
        index=pd.DatetimeIndex(dates, name="Date"),
        name="PX_LAST",
        dtype="float64",
    )


def _assert_hp_headers(ws: Worksheet) -> None:
    """Cheap belt-and-braces against accidental row edits before save."""
    if ws["B5"].value != "Date" or ws["C5"].value != "PX_LAST":
        raise WorkbookExtractionError(
            ws.title,
            f"header mismatch at row 5: B5={ws['B5'].value!r}, "
            f"C5={ws['C5'].value!r}",
        )


def _extract_hp_monthly(wb) -> PriceHistorySheet:
    """Extract the monthly close-price array spill (PR #65)."""
    ws = wb["HP_Monthly"]
    _assert_hp_headers(ws)
    na_cells: list[str] = []
    series = _read_price_series(ws, na_cells=na_cells)
    return PriceHistorySheet(series=series, na_cells=sorted(na_cells))


def _extract_hp_daily(wb) -> PriceHistorySheet:
    """Extract the daily close-price array spill (PR #65)."""
    ws = wb["HP_Daily"]
    _assert_hp_headers(ws)
    na_cells: list[str] = []
    series = _read_price_series(ws, na_cells=na_cells)
    return PriceHistorySheet(series=series, na_cells=sorted(na_cells))


def _extract_rv_comps(wb) -> RVCompsSheet | None:
    """STUB — populated in PR #65.

    Returns None when sheet absent (pre-lab Master template path).
    Returns empty RVCompsSheet when sheet present but extraction
    not yet wired.
    """
    if "RV_Comps" not in wb.sheetnames:
        return None
    return RVCompsSheet()


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> int:
    """Thin CLI shell. Usage: python -m mfip.ingestion.bloomberg.parser <path>"""
    cli = argparse.ArgumentParser(description=__doc__)
    cli.add_argument("path", type=Path, help="Path to Bloomberg workbook")
    cli.add_argument(
        "--policy",
        choices=[p.value for p in ValidationPolicy],
        default=ValidationPolicy.STRICT.value,
        help="Validation policy (default: strict)",
    )
    args = cli.parse_args()

    policy = ValidationPolicy(args.policy)
    try:
        result = parse_workbook(args.path, validation_policy=policy)
    except WorkbookValidationError as exc:
        print("FAIL: validation refused", file=sys.stderr)
        print(exc.report.render(), file=sys.stderr)
        return 1
    except BloombergParserError as exc:
        print(f"FAIL: {exc}", file=sys.stderr)
        return 1

    print(f"PARSED: {result.ticker_short}")
    print(f"  Validation: {result.validation_report.tier.value}")
    print(f"  Advisories: {len(result.advisories)}")
    print(f"  Source: {result.source_path}")
    return 0


if __name__ == "__main__":
    # Configure root logger only when invoked from CLI — never when
    # parse_workbook is called programmatically from another module.
    logging.basicConfig(
        level=logging.INFO,
        format="%(levelname)s %(name)s: %(message)s",
    )
    sys.exit(main())
